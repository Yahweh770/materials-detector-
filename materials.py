"""
Приложение для учёта материалов — Единая база
Версия: 2.0

Функционал:
- Добавление, редактирование и удаление записей о материалах
- Поиск по всем полям записи
- Экспорт данных в Excel с форматированием
- Отслеживание просроченных документов
- Работа с несколькими файлами баз данных
- Контекстное меню для быстрого доступа к функциям
- Копирование данных в буфер обмена

Автор: [Ваше имя]
Дата создания: 2024
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook


class MaterialApp:
    """Основной класс приложения для учёта материалов"""
    
    def __init__(self, root):
        """Инициализация приложения"""
        self.root = root
        self.root.title("Учёт материалов — Единая база")
        self.root.geometry("1920x980")
        self.root.minsize(1400, 700)
        self.root.state('zoomed')  # Запуск в полноэкранном режиме по умолчанию

        self.data = []
        self.next_id = 0
        self.today = datetime.now().date()
        self.data_file = self.get_data_file()
        self.backup_data = None  # Для временного хранения данных при фильтрации
        self._context_menu_event = None  # Для хранения события контекстного меню

        self.load_data()
        
        # Инициализация next_id после загрузки данных
        if self.data:
            self.next_id = max(item.get('id', 0) for item in self.data) + 1

        # ==================== МЕНЮ СВЕРХУ ====================
        menubar = tk.Menu(root)
        root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Сменить файл базы", command=self.change_data_file)
        file_menu.add_command(label="Экспорт в Excel", command=self.export_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.on_close)

        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Правка", menu=edit_menu)
        edit_menu.add_command(label="Добавить новую строку", command=self.add_new_row)
        edit_menu.add_command(label="Добавить новое поле", command=self.add_new_column)
        edit_menu.add_separator()
        edit_menu.add_command(label="Удалить строку", command=self.delete_row)

        # ==================== ПАНЕЛЬ ИНСТРУМЕНТОВ ====================
        toolbar = tk.Frame(root, relief="raised", bd=1)
        toolbar.pack(fill="x", padx=5, pady=5)

        tk.Button(toolbar, text="➕ Добавить материал", width=20, height=2, bg="#4CAF50", fg="white",
                  font=("Arial", 10, "bold"), command=self.open_add_material_window).pack(side="left", padx=5)
        
        tk.Button(toolbar, text="✏️ Редактировать", width=18, height=2, bg="#FF9800", fg="white",
                  command=self.edit_selected_row).pack(side="left", padx=5)
        
        tk.Button(toolbar, text="🗑️ Удалить", width=18, height=2, bg="#f44336", fg="white",
                  command=self.delete_row).pack(side="left", padx=5)
        
        tk.Button(toolbar, text="📊 Экспорт в Excel", width=18, height=2, bg="#2196F3", fg="white",
                  command=self.export_to_excel).pack(side="left", padx=5)
        
        tk.Button(toolbar, text="🔄 Сменить базу", width=18, height=2, bg="#9C27B0", fg="white",
                  command=self.change_data_file).pack(side="left", padx=5)

        tk.Label(toolbar, text="   ").pack(side="left")  # отступ

        # Поиск
        tk.Label(toolbar, text="Поиск:", font=("Arial", 10)).pack(side="left", padx=(20, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.search_data())
        self.search_entry = tk.Entry(toolbar, textvariable=self.search_var, width=30, font=("Arial", 10))
        self.search_entry.pack(side="left", padx=5)
        
        tk.Button(toolbar, text="🔍 Очистить", width=10, command=self.clear_search).pack(side="left", padx=5)
        
        # Контекстное меню для Treeview
        self.tree_context_menu = tk.Menu(self.root, tearoff=0)
        self.tree_context_menu.add_command(label="Копировать строку", command=self.copy_row_to_clipboard)
        self.tree_context_menu.add_command(label="Копировать ячейку", command=lambda: self.copy_cell_to_clipboard(None))
        self.tree_context_menu.add_separator()
        self.tree_context_menu.add_command(label="Редактировать", command=self.edit_selected_row)
        self.tree_context_menu.add_command(label="Удалить", command=self.delete_row)

        self.file_label = tk.Label(toolbar, text=f"Файл: {os.path.basename(self.data_file)}", 
                                  font=("Arial", 10, "bold"), fg="blue")
        self.file_label.pack(side="right", padx=15)

        # ==================== ТАБЛИЦА ====================
        self.tree = ttk.Treeview(root, show="headings", height=25)
        self.refresh_columns()

        scrollbar = ttk.Scrollbar(root, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True, padx=10, pady=5)
        scrollbar.pack(side="right", fill="y", padx=(0,10), pady=5)

        # Привязки для Treeview
        self.tree.bind("<Double-1>", self.on_double_click)
        self.tree.bind("<Delete>", self.delete_row)
        self.tree.bind("<Control-c>", self.copy_to_clipboard)
        
        # Контекстное меню для Treeview
        self.tree.bind("<Button-3>", self.show_tree_context_menu)
        
        # Глобальные привязки для работы с буфером обмена (Ctrl+C, Ctrl+V, Ctrl+X)
        self.root.bind("<Control-c>", self.global_copy)
        self.root.bind("<Control-v>", self.global_paste)
        self.root.bind("<Control-x>", self.global_cut)
        
        # ==================== НИЖНЯЯ ПАНЕЛЬ ====================
        bottom_frame = tk.Frame(root, relief="sunken", bd=1)
        bottom_frame.pack(fill="x", padx=10, pady=8)

        tk.Button(bottom_frame, text="💾 Сохранить все изменения", width=25, height=2,
                 bg="#2E7D32", fg="white", font=("Arial", 10, "bold"),
                 command=lambda: self.save_data(show_msg=True)).pack(side="left", padx=10)

        # Индикатор просроченных документов
        self.expired_label = tk.Label(bottom_frame, text="", font=("Arial", 10, "bold"), fg="red")
        self.expired_label.pack(side="left", padx=20)

        tk.Button(bottom_frame, text="⚠️ Просроченные документы", width=25, height=2,
                 bg="#FF5722", fg="white", font=("Arial", 10, "bold"),
                 command=self.show_expired_documents).pack(side="left", padx=10)

        tk.Button(bottom_frame, text="🔄 Показать все", width=15, height=2,
                 command=self.show_all_documents).pack(side="left", padx=5)

        tk.Button(bottom_frame, text="🚪 Выход", width=15, height=2, bg="#f44336", fg="white",
                 command=self.on_close).pack(side="right", padx=10)

        self.refresh_tree()
        
        # Обновление информации о просроченных документах
        self.update_expired_info()

    def show_tree_context_menu(self, event):
        """Показать контекстное меню для Treeview"""
        # Выбираем элемент под курсором
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
        
        # Сохраняем координаты события для использования в copy_cell_to_clipboard
        self._context_menu_event = event
        
        # Показываем меню
        self.tree_context_menu.tk_popup(event.x_root, event.y_root)

    def copy_row_to_clipboard(self):
        """Копирование всей строки в буфер обмена"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Внимание", "Выберите строку для копирования!")
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        cols = self.tree["columns"]
        
        # Формируем строку с заголовками
        header_row = '\t'.join(str(col) for col in cols)
        data_row = '\t'.join(str(v) for v in values)
        text_to_copy = f"{header_row}\n{data_row}"
        
        self.root.clipboard_clear()
        self.root.clipboard_append(text_to_copy)
        messagebox.showinfo("Успешно", "Строка скопирована в буфер обмена!")

    def copy_cell_to_clipboard(self, event=None):
        """Копирование содержимого конкретной ячейки в буфер обмена"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Внимание", "Выберите строку!")
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        
        # Пытаемся получить координаты из сохраненного события контекстного меню
        selected_value = None
        evt = event if event else getattr(self, '_context_menu_event', None)
        
        if evt and hasattr(evt, 'x') and hasattr(evt, 'y'):
            # Определяем колонку по координатам клика
            region = self.tree.identify_region(evt.x, evt.y)
            if region == 'cell':
                col_id = self.tree.identify_column(evt.x)
                if col_id:
                    # Получаем индекс колонки (формат: #N где N - номер)
                    try:
                        col_index = int(col_id[1:]) - 1
                        if 0 <= col_index < len(values):
                            selected_value = values[col_index]
                    except (ValueError, IndexError):
                        pass
        
        # Если не удалось определить конкретную ячейку, копируем первое значение
        if selected_value is None and values:
            selected_value = values[0]
        
        if selected_value is not None:
            self.root.clipboard_clear()
            self.root.clipboard_append(str(selected_value))
            messagebox.showinfo("Успешно", f"Ячейка скопирована: {selected_value}")
        else:
            messagebox.showwarning("Внимание", "Нет данных для копирования!")

    def search_data(self):
        """Поиск данных по введенному тексту"""
        search_text = self.search_var.get().lower().strip()
        
        if not search_text:
            self.refresh_tree()
            return
        
        # Очищаем текущее отображение
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Ищем совпадения во всех полях
        for row in self.data:
            match = False
            for key, value in row.items():
                if str(value).lower().find(search_text) >= 0:
                    match = True
                    break
            
            if match:
                values = [row.get(col, "") for col in self.tree["columns"]]
                self.tree.insert("", "end", values=values)
        
        # Обновляем информацию о количестве найденных записей
        found_count = self.tree.get_children()
        if len(found_count) == 0:
            self.expired_label.config(text=f"🔍 Ничего не найдено по запросу: {search_text}")
        else:
            self.expired_label.config(text=f"🔍 Найдено записей: {len(found_count)}")

    def clear_search(self):
        """Очистка поиска и отображение всех данных"""
        self.search_var.set("")
        self.refresh_tree()
        self.update_expired_info()

    # ==================== ОКНО ДОБАВЛЕНИЯ НОВОГО МАТЕРИАЛА ====================
    def open_add_material_window(self):
        win = tk.Toplevel(self.root)
        win.title("Добавление нового материала")
        win.geometry("700x600")
        win.resizable(True, True)

        # Основной фрейм для полей с прокруткой
        canvas = tk.Canvas(win)
        scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        fields = [
            ("Производитель *", "manufacturer"),
            ("Вид материала *", "material_type"),
            ("Паспорт №", "passport_num"),
            ("Дата производства", "production_date"),
            ("Срок хранения", "shelf_life"),
            ("Сертификат №", "cert_num"),
            ("Дата выдачи сертификата", "cert_issue_date"),
            ("Дата окончания сертификата", "cert_exp_date"),
            ("Протокол №", "lab_protocol_num"),
            ("Дата протокола", "lab_protocol_date"),
            ("Акт отбора №", "sample_act_num"),
            ("Дата акта", "sample_act_date"),
        ]

        entries = {}
        for i, (label_text, key) in enumerate(fields):
            tk.Label(scrollable_frame, text=label_text + ":", font=("Arial", 10)).grid(row=i, column=0, sticky="e", padx=10, pady=6)
            entry = tk.Entry(scrollable_frame, width=60, font=("Arial", 10))
            entry.grid(row=i, column=1, padx=10, pady=6)
            entries[key] = entry

        # Фрейм для дополнительных полей
        extra_fields_frame = tk.Frame(scrollable_frame)
        extra_fields_frame.grid(row=len(fields), column=0, columnspan=2, sticky="w", padx=10, pady=5)
        
        def add_extra_field():
            name = simpledialog.askstring("Дополнительное поле", 
                "Введите название дополнительного поля:")
            if not name:
                return
            key = name.strip().lower().replace(" ", "_").replace(":", "_")
            
            if key in entries:
                messagebox.showwarning("Ошибка", "Такое поле уже существует!")
                return

            row_idx = len(entries)
            tk.Label(extra_fields_frame, text=name + ":", font=("Arial", 10)).grid(row=row_idx-len(fields), column=0, sticky="e", padx=10, pady=6)
            entry = tk.Entry(extra_fields_frame, width=60, font=("Arial", 10))
            entry.grid(row=row_idx-len(fields), column=1, padx=10, pady=6)
            entries[key] = entry

        tk.Button(scrollable_frame, text="➕ Добавить поле", bg="#FF9800", fg="white",
                 command=add_extra_field).grid(row=len(fields)+1, column=0, columnspan=2, pady=10)

        def save_material():
            # Проверка заполнения обязательных полей
            manufacturer = entries["manufacturer"].get().strip()
            material_type = entries["material_type"].get().strip()
            
            if not manufacturer or not material_type:
                messagebox.showwarning("Внимание", "Поля 'Производитель' и 'Вид материала' обязательны для заполнения!")
                return
            
            new_item = {"id": self.next_id}
            self.next_id += 1

            for key, entry in entries.items():
                new_item[key] = entry.get().strip()

            self.data.append(new_item)
            self.refresh_columns()
            self.refresh_tree()
            self.save_data()
            win.destroy()
            messagebox.showinfo("Успешно", "Новый материал добавлен!")

        tk.Button(win, text="✅ Сохранить материал", width=20, height=2, bg="#4CAF50", fg="white",
                 command=save_material).pack(pady=10)

    def add_new_column_from_window(self):
        name = simpledialog.askstring("Новое поле", 
            "Введите название нового поля\n(например: Партия, Количество, Примечание, Поставщик)")
        if not name:
            return
        key = name.strip().lower().replace(" ", "_").replace(":", "_")
        
        if key in self.tree["columns"]:
            messagebox.showwarning("Ошибка", "Такое поле уже существует!")
            return

        for item in self.data:
            item[key] = ""

        self.refresh_columns()
        self.refresh_tree()
        messagebox.showinfo("Добавлено", f"Поле «{name}» добавлено во все записи.")

    # ==================== Остальные методы ====================
    def get_data_file(self):
        config_file = "config.json"
        if os.path.exists(config_file):
            try:
                with open(config_file, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    if config.get("data_file") and os.path.exists(config["data_file"]):
                        return config["data_file"]
            except:
                pass

        # Сначала пробуем использовать файл по умолчанию в текущей директории
        default_file = "materials.json"
        if os.path.exists(default_file):
            with open(config_file, "w", encoding="utf-8") as f:
                json.dump({"data_file": default_file}, f, ensure_ascii=False, indent=2)
            return default_file
        
        # Если файл не найден, предлагаем пользователю выбрать папку для создания нового файла
        folder_path = filedialog.askdirectory(title="Выберите папку для создания базы данных материалов")
        if not folder_path:
            # Если пользователь отменил выбор, используем файл по умолчанию в текущей директории
            file_path = default_file
        else:
            file_path = os.path.join(folder_path, "materials.json")
        
        # Создаём новый пустой файл базы данных
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump([], f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать файл базы данных: {e}")
            self.root.destroy()
            exit()

        with open(config_file, "w", encoding="utf-8") as f:
            json.dump({"data_file": file_path}, f, ensure_ascii=False, indent=2)
        
        messagebox.showinfo("Успешно", f"Файл базы данных создан:\n{file_path}")
        return file_path

    def load_data(self):
        """Загрузка данных из JSON файла"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось загрузить данные: {e}")
                self.data = []
        else:
            self.data = []

    def save_data(self, show_msg=False):
        """Сохранение данных в JSON файл"""
        try:
            # Очищаем временные поля перед сохранением
            data_to_save = []
            for item in self.data:
                item_copy = {k: v for k, v in item.items() if not k.startswith('_')}
                data_to_save.append(item_copy)
            
            with open(self.data_file, "w", encoding="utf-8") as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=2)
            if show_msg:
                messagebox.showinfo("Успешно", "Данные успешно сохранены!")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {e}")

    def refresh_tree(self):
        """Обновление таблицы с данными"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for row in self.data:
            values = [row.get(col, "") for col in self.tree["columns"]]
            self.tree.insert("", "end", values=values)

    def refresh_columns(self):
        if not self.data:
            cols = ["manufacturer", "material_type", "cert_exp_date"]
        else:
            cols = [k for k in self.data[0].keys() if k != "id"]

        self.tree["columns"] = cols
        russian_names = {
            "manufacturer": "Производитель", "material_type": "Вид материала",
            "passport_num": "Паспорт №", "production_date": "Дата производства",
            "shelf_life": "Срок хранения", "cert_num": "Сертификат №",
            "cert_issue_date": "Дата выдачи", "cert_exp_date": "Дата окончания сертификата",
            "lab_protocol_num": "Протокол №", "lab_protocol_date": "Дата протокола",
            "sample_act_num": "Акт отбора №", "sample_act_date": "Дата акта"
        }
        for col in cols:
            name = russian_names.get(col, col.replace("_", " ").title())
            self.tree.heading(col, text=name)
            self.tree.column(col, width=160, anchor="center")

    def export_to_excel(self):
        """Экспорт данных в Excel файл с форматированием"""
        if not self.data:
            messagebox.showwarning("Внимание", "Нет данных для экспорта!")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Сохранить Excel файл",
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx")],
            initialfile=f"materials_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Материалы"
            
            # Получаем все колонки из данных (включая id)
            all_keys = set()
            for item in self.data:
                all_keys.update(item.keys())
            cols = ["id"] + [k for k in self.tree["columns"] if k != "id"]
            
            # Заголовки с русскими названиями
            russian_names = {
                "id": "№ п/п",
                "manufacturer": "Производитель", 
                "material_type": "Вид материала",
                "passport_num": "Паспорт №", 
                "production_date": "Дата производства",
                "shelf_life": "Срок хранения", 
                "cert_num": "Сертификат №",
                "cert_issue_date": "Дата выдачи сертификата", 
                "cert_exp_date": "Дата окончания сертификата",
                "lab_protocol_num": "Протокол №", 
                "lab_protocol_date": "Дата протокола",
                "sample_act_num": "Акт отбора №", 
                "sample_act_date": "Дата акта"
            }
            
            # Стили для заголовка
            header_font = openpyxl.styles.Font(bold=True, size=11, name="Arial")
            header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                bottom=openpyxl.styles.Side(style="thin")
            )
            
            # Стили для ячеек
            cell_alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
            cell_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                bottom=openpyxl.styles.Side(style="thin")
            )
            
            # Записываем заголовки
            for col_idx, col in enumerate(cols, 1):
                header = russian_names.get(col, col.replace("_", " ").title())
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            # Записываем данные с нумерацией строк
            for row_idx, row in enumerate(self.data, 2):
                for col_idx, col in enumerate(cols, 1):
                    value = row.get(col, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cell_alignment
                    cell.border = cell_border
            
            # Автоширина колонок
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(max(adjusted_width, 15), 60)
            
            # Замораживаем первую строку (заголовок)
            ws.freeze_panes = "A2"
            
            # Добавляем информацию об экспорте в конце
            total_rows = len(self.data)
            info_row = len(self.data) + 3
            ws.cell(row=info_row, column=1, value=f"Всего записей: {total_rows}")
            ws.cell(row=info_row+1, column=1, value=f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            
            wb.save(file_path)
            messagebox.showinfo("Успешно", f"Данные успешно экспортированы в:\n{file_path}\n\nЭкспортировано записей: {total_rows}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать данные:\n{e}")

    def change_data_file(self):
        """Смена файла базы данных"""
        file_path = filedialog.askopenfilename(
            title="Выберите новый файл базы данных",
            filetypes=[("JSON файл", "*.json")]
        )
        if file_path:
            config_file = "config.json"
            with open(config_file, "w", encoding="utf-8") as f:
                json.dump({"data_file": file_path}, f, ensure_ascii=False, indent=2)
            self.data_file = file_path
            self.load_data()
            if self.data:
                self.next_id = max(item.get('id', 0) for item in self.data) + 1
            else:
                self.next_id = 0
            self.refresh_columns()
            self.refresh_tree()
            self.file_label.config(text=f"Файл: {os.path.basename(self.data_file)}")
            messagebox.showinfo("Успешно", "База данных успешно изменена!")

    def add_new_row(self):
        """Добавление новой строки через меню"""
        self.open_add_material_window()

    def add_new_column(self):
        """Добавление нового поля через меню"""
        self.add_new_column_from_window()

    def edit_selected_row(self):
        """Редактирование выбранной строки"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Внимание", "Выберите строку для редактирования!")
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        
        # Находим соответствующую запись в данных
        cols = self.tree["columns"]
        record = {}
        for i, col in enumerate(cols):
            if i < len(values):
                record[col] = values[i]
        
        # Добавляем ID если есть
        for row in self.data:
            match = True
            for col in cols:
                if row.get(col, "") != record.get(col, ""):
                    match = False
                    break
            if match:
                record['id'] = row.get('id')
                break
        
        self.edit_record(record)

    def delete_row(self, event=None):
        """Удаление выбранной строки"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Внимание", "Выберите строку для удаления!")
            return
        
        if not messagebox.askyesno("Подтверждение", "Вы действительно хотите удалить выбранную строку?"):
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        
        # Находим соответствующую запись в данных
        cols = self.tree["columns"]
        record = {}
        for i, col in enumerate(cols):
            if i < len(values):
                record[col] = values[i]
        
        # Удаляем из данных
        for i, row in enumerate(self.data):
            match = True
            for col in cols:
                if row.get(col, "") != record.get(col, ""):
                    match = False
                    break
            if match:
                del self.data[i]
                break
        
        self.refresh_tree()
        self.save_data()
        messagebox.showinfo("Успешно", "Строка удалена!")

    def on_double_click(self, event):
        """Обработка двойного клика по строке"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        
        # Находим соответствующую запись в данных
        cols = self.tree["columns"]
        record = {}
        for i, col in enumerate(cols):
            if i < len(values):
                record[col] = values[i]
        
        # Добавляем ID если есть
        for row in self.data:
            match = True
            for col in cols:
                if row.get(col, "") != record.get(col, ""):
                    match = False
                    break
            if match:
                record['id'] = row.get('id')
                break
        
        self.edit_record(record)

    def edit_record(self, record):
        """Редактирование записи"""
        win = tk.Toplevel(self.root)
        win.title("Редактирование материала")
        win.geometry("700x600")
        win.resizable(True, True)

        fields = [
            ("Производитель", "manufacturer"),
            ("Вид материала", "material_type"),
            ("Паспорт №", "passport_num"),
            ("Дата производства", "production_date"),
            ("Срок хранения", "shelf_life"),
            ("Сертификат №", "cert_num"),
            ("Дата выдачи сертификата", "cert_issue_date"),
            ("Дата окончания сертификата", "cert_exp_date"),
            ("Протокол №", "lab_protocol_num"),
            ("Дата протокола", "lab_protocol_date"),
            ("Акт отбора №", "sample_act_num"),
            ("Дата акта", "sample_act_date"),
        ]

        entries = {}
        for i, (label_text, key) in enumerate(fields):
            tk.Label(win, text=label_text + ":", font=("Arial", 10)).grid(row=i, column=0, sticky="e", padx=10, pady=6)
            entry = tk.Entry(win, width=60, font=("Arial", 10))
            entry.grid(row=i, column=1, padx=10, pady=6)
            entry.insert(0, record.get(key, ""))
            entries[key] = entry
            
            # Добавляем локальные привязки для копирования/вставки
            entry.bind("<Control-c>", lambda e: e.widget.event_generate('<<Copy>>'))
            entry.bind("<Control-v>", lambda e: e.widget.event_generate('<<Paste>>'))
            entry.bind("<Control-x>", lambda e: e.widget.event_generate('<<Cut>>'))
            
            # Добавляем контекстное меню по правой кнопке
            entry.bind("<Button-3>", lambda e: self.show_context_menu(e))

        # Добавляем дополнительные поля
        extra_fields = [k for k in self.tree["columns"] if k not in [f[1] for f in fields]]
        for key in extra_fields:
            i = len(entries)
            tk.Label(win, text=key.replace("_", " ").title() + ":", font=("Arial", 10)).grid(row=i, column=0, sticky="e", padx=10, pady=6)
            entry = tk.Entry(win, width=60, font=("Arial", 10))
            entry.grid(row=i, column=1, padx=10, pady=6)
            entry.insert(0, record.get(key, ""))
            entries[key] = entry
            
            # Добавляем локальные привязки для копирования/вставки
            entry.bind("<Control-c>", lambda e: e.widget.event_generate('<<Copy>>'))
            entry.bind("<Control-v>", lambda e: e.widget.event_generate('<<Paste>>'))
            entry.bind("<Control-x>", lambda e: e.widget.event_generate('<<Cut>>'))
            entry.bind("<Button-3>", lambda e: self.show_context_menu(e))

        def save_changes():
            record_id = record.get('id')
            for row in self.data:
                if row.get('id') == record_id:
                    for key, entry in entries.items():
                        row[key] = entry.get().strip()
                    break
            
            self.refresh_tree()
            self.save_data()
            win.destroy()
            messagebox.showinfo("Успешно", "Запись обновлена!")

        tk.Button(win, text="💾 Сохранить изменения", width=20, height=2, bg="#4CAF50", fg="white",
                 command=save_changes).pack(pady=20)
                 
        # Делаем окно активным и передаем фокус первому полю
        win.focus_set()
        if entries:
            list(entries.values())[0].focus_set()
    
    def show_context_menu(self, event):
        """Показать контекстное меню для текстовых полей"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Копировать", command=lambda: event.widget.event_generate('<<Copy>>'))
        menu.add_command(label="Вставить", command=lambda: event.widget.event_generate('<<Paste>>'))
        menu.add_command(label="Вырезать", command=lambda: event.widget.event_generate('<<Cut>>'))
        menu.tk_popup(event.x_root, event.y_root)

    def copy_to_clipboard(self, event):
        """Копирование выделенной ячейки в Treeview"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        
        # Копируем первую выделенную ячейку (можно улучшить для конкретной ячейки)
        if values:
            self.root.clipboard_clear()
            self.root.clipboard_append(str(values[0]))

    def global_copy(self, event):
        """Глобальное копирование из любого текстового поля или Treeview"""
        # Проверяем, есть ли фокус на Entry или Text виджете
        focus_widget = self.root.focus_get()
        
        if focus_widget:
            widget_type = type(focus_widget).__name__
            
            # Если фокус на Entry или Text - используем стандартное копирование
            if widget_type in ['Entry', 'Text', 'Spinbox']:
                try:
                    focus_widget.event_generate('<<Copy>>')
                    return
                except:
                    pass
            
            # Если фокус на Treeview - копируем выделенную строку
            if isinstance(focus_widget, ttk.Treeview):
                selection = self.tree.selection()
                if selection:
                    item = self.tree.item(selection[0])
                    values = item['values']
                    if values:
                        # Копируем все значения строки, разделённые табуляцией
                        text_to_copy = '\t'.join(str(v) for v in values)
                        self.root.clipboard_clear()
                        self.root.clipboard_append(text_to_copy)
                        return

    def global_paste(self, event):
        """Глобальная вставка в любое текстовое поле или активную ячейку"""
        focus_widget = self.root.focus_get()
        
        if focus_widget:
            widget_type = type(focus_widget).__name__
            
            # Если фокус на Entry или Text - используем стандартную вставку
            if widget_type in ['Entry', 'Text', 'Spinbox']:
                try:
                    focus_widget.event_generate('<<Paste>>')
                    return
                except:
                    pass
            
            # Если фокус на Treeview - показываем сообщение
            if isinstance(focus_widget, ttk.Treeview):
                messagebox.showinfo("Инфо", 
                    "Для вставки данных в таблицу:\n"
                    "1. Дважды кликните на ячейку для редактирования\n"
                    "2. Используйте Ctrl+V в окне редактирования")
                return

    def global_cut(self, event):
        """Глобальное вырезание из любого текстового поля"""
        focus_widget = self.root.focus_get()
        
        if focus_widget:
            widget_type = type(focus_widget).__name__
            
            # Если фокус на Entry или Text - используем стандартное вырезание
            if widget_type in ['Entry', 'Text', 'Spinbox']:
                try:
                    focus_widget.event_generate('<<Cut>>')
                    return
                except:
                    pass

    def paste_from_clipboard(self, event):
        """Вставка из буфера обмена (заглушка - теперь используется global_paste)"""
        self.global_paste(event)

    def update_expired_info(self):
        """Обновление информации о просроченных документах"""
        expired_count = 0
        expiring_soon_count = 0
        
        for item in self.data:
            # Проверяем дату окончания сертификата
            cert_exp_date = item.get('cert_exp_date', '').strip()
            if cert_exp_date:
                try:
                    exp_date = datetime.strptime(cert_exp_date, '%d.%m.%Y').date()
                    days_left = (exp_date - self.today).days
                    
                    if days_left < 0:
                        expired_count += 1
                    elif days_left <= 30:
                        expiring_soon_count += 1
                except ValueError:
                    pass
            
            # Проверяем дату протокола (если есть срок действия)
            protocol_date = item.get('lab_protocol_date', '').strip()
            if protocol_date:
                try:
                    prot_date = datetime.strptime(protocol_date, '%d.%m.%Y').date()
                    # Протоколы обычно действительны 1 год
                    exp_date = prot_date + timedelta(days=365)
                    days_left = (exp_date - self.today).days
                    
                    if days_left < 0:
                        expired_count += 1
                    elif days_left <= 30:
                        expiring_soon_count += 1
                except ValueError:
                    pass
            
            # Проверяем дату акта отбора
            act_date = item.get('sample_act_date', '').strip()
            if act_date:
                try:
                    act_dt = datetime.strptime(act_date, '%d.%m.%Y').date()
                    # Акты обычно действительны 1 год
                    exp_date = act_dt + timedelta(days=365)
                    days_left = (exp_date - self.today).days
                    
                    if days_left < 0:
                        expired_count += 1
                    elif days_left <= 30:
                        expiring_soon_count += 1
                except ValueError:
                    pass
        
        # Обновляем метку с информацией
        if expired_count > 0 and expiring_soon_count > 0:
            self.expired_label.config(
                text=f"⚠️ Просрочено: {expired_count} | Истекает в течение 30 дней: {expiring_soon_count}"
            )
        elif expired_count > 0:
            self.expired_label.config(text=f"🔴 Просрочено документов: {expired_count}")
        elif expiring_soon_count > 0:
            self.expired_label.config(text=f"🟡 Истекает в течение 30 дней: {expiring_soon_count}")
        else:
            self.expired_label.config(text="✅ Все документы действительны")

    def show_expired_documents(self):
        """Отображение только просроченных и истекающих документов"""
        expired_items = []
        
        for item in self.data:
            is_expired = False
            is_expiring_soon = False
            expiry_info = []
            
            # Проверяем дату окончания сертификата
            cert_exp_date = item.get('cert_exp_date', '').strip()
            if cert_exp_date:
                try:
                    exp_date = datetime.strptime(cert_exp_date, '%d.%m.%Y').date()
                    days_left = (exp_date - self.today).days
                    
                    if days_left < 0:
                        is_expired = True
                        expiry_info.append(f"Сертификат просрочен на {-days_left} дн.")
                    elif days_left <= 30:
                        is_expiring_soon = True
                        expiry_info.append(f"Сертификат истекает через {days_left} дн.")
                except ValueError:
                    pass
            
            # Проверяем дату протокола
            protocol_date = item.get('lab_protocol_date', '').strip()
            if protocol_date:
                try:
                    prot_date = datetime.strptime(protocol_date, '%d.%m.%Y').date()
                    exp_date = prot_date + timedelta(days=365)
                    days_left = (exp_date - self.today).days
                    
                    if days_left < 0:
                        is_expired = True
                        expiry_info.append(f"Протокол просрочен на {-days_left} дн.")
                    elif days_left <= 30:
                        is_expiring_soon = True
                        expiry_info.append(f"Протокол истекает через {days_left} дн.")
                except ValueError:
                    pass
            
            # Проверяем дату акта отбора
            act_date = item.get('sample_act_date', '').strip()
            if act_date:
                try:
                    act_dt = datetime.strptime(act_date, '%d.%m.%Y').date()
                    exp_date = act_dt + timedelta(days=365)
                    days_left = (exp_date - self.today).days
                    
                    if days_left < 0:
                        is_expired = True
                        expiry_info.append(f"Акт просрочен на {-days_left} дн.")
                    elif days_left <= 30:
                        is_expiring_soon = True
                        expiry_info.append(f"Акт истекает через {days_left} дн.")
                except ValueError:
                    pass
            
            if is_expired or is_expiring_soon:
                item_copy = item.copy()
                item_copy['_expiry_info'] = '; '.join(expiry_info)
                expired_items.append(item_copy)
        
        if not expired_items:
            messagebox.showinfo("Инфо", "Нет просроченных или истекающих документов!")
            return
        
        # Сохраняем текущие данные для восстановления
        self.backup_data = self.data.copy()
        self.data = expired_items
        self.refresh_tree()
        
        # Добавляем колонку с информацией об истечении
        cols = list(self.tree["columns"])
        if '_expiry_info' not in cols:
            cols.append('_expiry_info')
            self.tree["columns"] = cols
            self.tree.heading('_expiry_info', text='⚠️ Информация о сроках')
            self.tree.column('_expiry_info', width=300, anchor='w')

    def show_all_documents(self):
        """Возврат к отображению всех документов"""
        if hasattr(self, 'backup_data') and self.backup_data:
            self.data = self.backup_data
            self.backup_data = None
            self.refresh_columns()
            self.refresh_tree()
            messagebox.showinfo("Инфо", "Отображаются все документы")
        else:
            self.refresh_columns()
            self.refresh_tree()

    def on_close(self):
        self.save_data()
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    
    # Установка иконки приложения (если доступна)
    try:
        root.iconbitmap(default='icon.ico')
    except:
        pass
    
    app = MaterialApp(root)
    root.mainloop()